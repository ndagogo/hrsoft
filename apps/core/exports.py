"""Shared helpers for CSV and Excel file downloads."""

import csv
import io
from datetime import datetime

from django.http import HttpResponse


def timestamped_filename(prefix: str, ext: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.{ext}"


def build_csv_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    response = HttpResponse(
        "\ufeff" + buffer.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_excel_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Export"

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_export_response(
    filename_prefix: str,
    fmt: str,
    headers: list[str],
    rows: list[list],
) -> HttpResponse:
    if fmt == "xlsx":
        return build_excel_response(timestamped_filename(filename_prefix, "xlsx"), headers, rows)
    return build_csv_response(timestamped_filename(filename_prefix, "csv"), headers, rows)


def read_uploaded_rows(uploaded_file) -> tuple[list[str], list[dict]]:
    """Return (headers, row dicts) from an uploaded CSV or Excel file."""
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(uploaded_file)
    if name.endswith(".csv"):
        return _read_csv(uploaded_file)
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def _read_csv(uploaded_file) -> tuple[list[str], list[dict]]:
    raw = uploaded_file.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig")
    else:
        text = raw
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("The CSV file has no header row.")
    headers = [h.strip() for h in reader.fieldnames]
    rows = [{k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()} for row in reader]
    return headers, rows


def _read_xlsx(uploaded_file) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("The Excel file is empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if not any(headers):
        raise ValueError("The Excel file has no header row.")

    rows = []
    for values in rows_iter:
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = values[idx] if idx < len(values) else None
            if value is None:
                row[header] = ""
            elif isinstance(value, str):
                row[header] = value.strip()
            else:
                row[header] = value
        rows.append(row)
    return headers, rows
